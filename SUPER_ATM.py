#! /usr/bin/env python

#SUPERimposed Automated Trace Maker = SUPER_ATM
#Version 1.0
#Released to Github on 4/4/2021.

#This Python script will take .CSV HPLC files (containing only retention times and
#absorbance values) and create superimposed unnormalized and normalized Excel plots.

#The following changes the working directory to the folder in which the Python executable
#is stored in. Disable this if you are not using the executable.
#os.chdir(os.path.dirname(sys.executable))

#Imports important modules and classes.
import csv
import codecs
import glob
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (Paragraph, ParagraphProperties, CharacterProperties, 
									Font)
from openpyxl.styles import Alignment
import sys
import re

#This allows for the .CSV files to be entered into SUPER_ATM via ascending order.
numbers = re.compile(r'(\d+)')
def numericalSort(value):
	"""Function for sorting files in numerical and alphabetical ascending order"""
	parts = numbers.split(value)
	parts[1::2] = map(int, parts[1::2])
	return parts

#Introduces user to the program.
print "Welcome to SUPER_ATM (SUPERimposed Automated Trace Maker)! This program will"
print "prepare superimposed HPLC traces of every .CSV file within a folder. The"
print ".CSV files only need to contain the retention times and absorbance values."
print ""

#Creates master lists to store information about each .CSV file.
minRTList = []
maxRTList = []
RTList = []
absorbanceList = []
normAbsList = []
fileNameList = []

#Enables user to enter the same retention times for all files in a folder.
sameRT = False
sameRTAns = raw_input("Do all .CSV files have the same retention times?")
print ""
if sameRTAns == "" or sameRTAns[0].lower() == "y":
	sameRT = True
	userInputs = False
	while userInputs == False:
		minRT = raw_input("Using only numbers, enter the minimum retention time: ")
		print ""
		maxRT = raw_input("Using only numbers, enter the maximum retention time: ")
		print ""
		try:
			minRT = float(minRT)
			maxRT = float(maxRT)
			print "You have entered the following retention times for all files:"
			print ""
			print "Minimum Retention Time = "+str(minRT)
			print "Maximum Retention Time = "+str(maxRT)
			print ""
			correctIn = raw_input("Are these values correct?")
			print ""
			if correctIn == "" or correctIn[0].lower() == "y":
				if minRT < maxRT:
					userInputs = True
				else:
					print "The minimum retention time cannot be larger than or equal to"
					print "the maximum retention time!"
					print "Please enter the retention times again."
					print ""
		except ValueError:
			print "Only numbers can be entered for the retention times!"
			print "Please enter the retention times again."
			print ""

#Enables user to stagger superimposed traces by shifting values on the x- and/or y-axis.
xStag = False
xShift = raw_input("Would you like to stagger the unnormalized trace by shifting x-axis "
					"values?")
print ""
if xShift == "" or xShift[0].lower() == "y":
	xStag = True
	print "After the first .CSV file is plotted, the remaining .CSV files will be shifted"
	print "by the specific number of retention time minutes you enter."
	print ""
	xShiftInput = False
	while xShiftInput == False:
		RTShift = raw_input("Using only numbers, enter the retention time shift: ")
		print ""
		try:
			RTShift = float(RTShift)
			print "You have entered the following retention time shift:"
			print ""
			print str(RTShift)+" min"
			print ""
			correctIn = raw_input("Is this retention time shift correct?")
			print ""
			if correctIn == "" or correctIn[0].lower() == "y":
				xShiftInput = True
		except ValueError:
			print "Only numbers can be entered for the retention time shift!"
			print "Please enter the retention time shift again."
			print ""
yStag = False
yShift = raw_input("Would you like to stagger the unnormalized trace by shifting y-axis "
					"values?")
print ""
if yShift == "" or yShift[0].lower() == "y":
	yStag = True
	print "After the first .CSV file is plotted, the remaining .CSV files will be shifted"
	print "by the specific mAU value you enter."
	print ""
	yShiftInput = False
	while yShiftInput == False:
		absShift = raw_input("Using only numbers, enter the mAU shift: ")
		print ""
		try:
			absShift = float(absShift)
			print "You have entered the following mAU shift:"
			print ""
			print str(absShift)+" mAU"
			print ""
			correctIn = raw_input("Is this mAU shift correct?")
			print ""
			if correctIn == "" or correctIn[0].lower() == "y":
				yShiftInput = True
		except ValueError:
			print "Only numbers can be entered for the mAU shift!"
			print "Please enter the mAU shift again."
			print ""

#Performs the following loop on each .CSV file within a folder.
for File in sorted(glob.iglob("*.[Cc][Ss][Vv]"), key=numericalSort):
    #Saves the file name in the master list.
    fileNameList.append(File[:-4])
    
    #Tells user which file is currently being prepared for plotting.
    print File+" is being prepared for plotting."
    print ""
    
    #Asks user for minimum and maximum retention times that are desired to be plotted
    #and ensures the inputted values are numerical.
    userInputs = False
    while userInputs == False and sameRT == False:
    	minRT = raw_input("Using only numbers, enter the minimum retention time: ")
    	print ""
    	maxRT = raw_input("Using only numbers, enter the maximum retention time: ")
    	print ""
    	try:
    		minRT = float(minRT)
    		maxRT = float(maxRT)
    		print "You have entered the following retention times for"
    		print File+":"
    		print ""
    		print "Minimum Retention Time = "+str(minRT)
    		print "Maximum Retention Time = "+str(maxRT)
    		print ""
    		correctIn = raw_input("Are these values correct?")
    		print ""
    		if correctIn == "" or correctIn[0].lower() == "y":
    			if minRT < maxRT:
    				userInputs = True
    			else:
    				print "The minimum retention time cannot be larger than or equal to"
    				print "the maximum retention time!"
    				print "Please enter the retention times again."
    				print ""
    	except ValueError:
    		print "Only numbers can be entered for the retention times!"
    		print "Please enter the retention times again."
    		print ""
    
    #Stores min and max retention times in the master list.
    minRTList.append(minRT)
    maxRTList.append(maxRT)
    
    #Puts the retention times and absorbances into separate lists, based on user's inputs.
    retentionTimes = []
    absorbances = []
    inFile = codecs.open(File, "rU", "utf-16") #codecs used to convert .CSV properly.
    csvReader = csv.reader(inFile)
    for row in csvReader:
    	formattedRow = row[0].split("\t")
    	if float(formattedRow[0]) >= minRT and float(formattedRow[0]) <= maxRT:
    		retentionTimes.append(float(formattedRow[0]))
    		absorbances.append(float(formattedRow[1]))
    inFile.close()
    
    #Stores retention times and absorbances into the master lists.
    RTList.append(retentionTimes)
    absorbanceList.append(absorbances)
    
    #Finds lowest absorbance value to eventually normalize all data. Also makes sure that
    #user did not enter incorrect retention time information or have an empty CSV file,
    #which would result in an empty absorbance list.
    try:
    	minAbs = min(absorbances)
    except ValueError:
    	print "ERROR! There is no data to plot!"
    	print "You have either entered an empty CSV file, or you have entered the wrong" 
    	print "retention times for this file."
    	print ""
    	print "Please restart the program with the proper file or retention times."
    	exit = raw_input("Press 'enter' to close the program.")
    	sys.exit()
    
    #Sets the lowest absorbance value as zero and corrects all absorbances based on this
    #zero value.
    balancedAbs = []
    for abs in absorbances:
    	balancedAbs.append(abs-minAbs)
    	
    #Finds the maximum balanced absorbance and normalizes all absorbances with this value.
    maxAbs = max(balancedAbs)
    normAbs = []
    for abs in balancedAbs:
    	normAbs.append(abs/maxAbs*100)
    
    #Stores the normalized absorbances in the master list.
    normAbsList.append(normAbs)
    
#Tells user that the superimposed plots are now being created.
print "Preparing superimposed plots..."
print ""
        	
#Sets up the output Excel file.
outFile = openpyxl.Workbook()
outFile.create_sheet(index=-1, title="Superimposed Plots")
outFile.create_sheet(index=-1, title="Raw Data")
if xStag == True or yStag == True:
	outFile.create_sheet(index=-1, title="Staggered Data")
outFile.remove_sheet(outFile.get_sheet_by_name("Sheet"))
sheet = outFile.get_sheet_by_name("Raw Data")

#Sets up the superimposed plot charts.
rawChart = ScatterChart()
rawChart.y_axis.title = "Absorbance (mAU)"
rawChart.x_axis.title = "Retention Time (min)"
normChart = ScatterChart()
normChart.y_axis.title = "Normalized Absorbance"
normChart.x_axis.title = "Retention Time (min)"

#Sets the chart title text font and size.
titleFont = Font(typeface="Times New Roman")
textProp = CharacterProperties(latin=titleFont, sz=1200)
rawChart.x_axis.title.tx.rich.p[0].r.rPr = textProp
rawChart.y_axis.title.tx.rich.p[0].r.rPr = textProp
normChart.x_axis.title.tx.rich.p[0].r.rPr = textProp
normChart.y_axis.title.tx.rich.p[0].r.rPr = textProp
    
#Sets the axis text font and size.
axisFont = Font(typeface="Times New Roman")
fontProp = CharacterProperties(latin=axisFont, sz=1000)
finalProp = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=fontProp), 
			endParaRPr=fontProp)])
rawChart.x_axis.txPr = finalProp
rawChart.y_axis.txPr = finalProp
normChart.x_axis.txPr = finalProp
normChart.y_axis.txPr = finalProp

#Places all data into the Excel file and creates the superimposed plots.
dataTracker = 0
columnNum = 1
for i in minRTList:
	#Merges top 3 cells to create a header for the .CSV file in the raw data sheet.
    sheet.merge_cells(start_row=1, start_column=columnNum, end_row=1, 
    					end_column=columnNum+2)
    cell = sheet.cell(row = 1, column = columnNum)
    cell.value = fileNameList[dataTracker]
    cell.alignment = Alignment(horizontal="center")
    
    #Merges top 3 cells to create a header for the .CSV file if staggered data is used.
    if xStag == True or yStag == True:
    	sheet = outFile.get_sheet_by_name("Staggered Data")
    	sheet.merge_cells(start_row=1, start_column=columnNum, end_row=1, 
    					end_column=columnNum+1)
    	cell = sheet.cell(row = 1, column = columnNum)
    	cell.value = fileNameList[dataTracker]
    	cell.alignment = Alignment(horizontal="center")
    	sheet = outFile.get_sheet_by_name("Raw Data")
    
    #Puts retention times into the Excel file.
    cell = sheet.cell(row = 2, column = columnNum)
    cell.value = "Retention Time (Min)"
    rowNum = 3
    for time in RTList[dataTracker]:
    	cell = sheet.cell(row = rowNum, column = columnNum)
    	cell.value = time
    	rowNum += 1
    
    #If staggering is desired, places shifted retention times into shifted sheet.
    if xStag == True or yStag == True:
    	sheet = outFile.get_sheet_by_name("Staggered Data")
    	cell = sheet.cell(row = 2, column = columnNum)
    	cell.value = "Retention Time (Min)"
    	rowNum = 3
    	for time in RTList[dataTracker]:
    		cell = sheet.cell(row = rowNum, column = columnNum)
    		if xStag == False or dataTracker == 0:
    			cell.value = time
    		else:
    			cell.value = time + (RTShift*dataTracker)
    		rowNum += 1
    	sheet = outFile.get_sheet_by_name("Raw Data")
    columnNum += 1
    	
    #Puts absorbances into the Excel file.
    cell = sheet.cell(row = 2, column = columnNum)
    cell.value = "Raw Absorbance (mAU)"
    rowNum = 3
    for abs in absorbanceList[dataTracker]:
    	cell = sheet.cell(row = rowNum, column = columnNum)
    	cell.value = abs
    	rowNum += 1
    
    #If staggering is desired, places shifted mAU values into shifted sheet.
    if xStag == True or yStag == True:
    	sheet = outFile.get_sheet_by_name("Staggered Data")
    	cell = sheet.cell(row = 2, column = columnNum)
    	cell.value = "Absorbance (mAU)"
    	rowNum = 3
    	for abs in absorbanceList[dataTracker]:
    		cell = sheet.cell(row = rowNum, column = columnNum)
    		if yStag == False or dataTracker == 0:
    			cell.value = abs
    		else:
    			cell.value = abs + (absShift*dataTracker)
    		rowNum += 1
    columnNum += 1
    
    #Adds data series to raw chart.
    if xStag == True or yStag == True:
    	sheet = outFile.get_sheet_by_name("Staggered Data")
    else:
    	sheet = outFile.get_sheet_by_name("Raw Data")
    xvalues = Reference(sheet, min_col=columnNum-2, min_row=3, max_row=rowNum-1)
    yvalues = Reference(sheet, min_col=columnNum-1, min_row=3, max_row=rowNum-1)
    series = Series(values = yvalues, xvalues = xvalues, title=fileNameList[dataTracker])
    series.graphicalProperties.line.width = 100
    rawChart.series.append(series)
    
    #Puts normalized absorbances into the Excel file.
    sheet = outFile.get_sheet_by_name("Raw Data")
    cell = sheet.cell(row = 2, column = columnNum)
    cell.value = "Normalized Absorbance"
    rowNum = 3
    for abs in normAbsList[dataTracker]:
    	cell = sheet.cell(row = rowNum, column = columnNum)
    	cell.value = abs
    	rowNum += 1
    
    #Adds data series to normalized chart.
    xvalues = Reference(sheet, min_col=columnNum-2, min_row=3, max_row=rowNum-1)
    yvalues = Reference(sheet, min_col=columnNum, min_row=3, max_row=rowNum-1)
    series = Series(values = yvalues, xvalues = xvalues, title=fileNameList[dataTracker])
    series.graphicalProperties.line.width = 100
    normChart.series.append(series)
    
    #Adds 1 to count tracker variables.
    dataTracker += 1
    columnNum += 1

#Plots the superimposed charts in the Excel file.
sheet = outFile.get_sheet_by_name("Superimposed Plots")
sheet.add_chart(rawChart, "B5")
sheet.add_chart(normChart, "K5")

#Formats both charts with no gridlines.
noGridLines = GraphicalProperties(ln=LineProperties(noFill=True))
rawChart.x_axis.majorGridlines.spPr = noGridLines
rawChart.y_axis.majorGridlines.spPr = noGridLines
normChart.x_axis.majorGridlines.spPr = noGridLines
normChart.y_axis.majorGridlines.spPr = noGridLines

#Sets the min and max values for the x-axis in the unnormalized chart.
if xStag == True and RTShift < 0:
	rawChart.x_axis.scaling.min = min(minRTList) + (RTShift*(dataTracker-1))
	rawChart.x_axis.scaling.max = max(maxRTList)
elif xStag == True and RTShift > 0:
	rawChart.x_axis.scaling.min = min(minRTList)
	rawChart.x_axis.scaling.max = max(maxRTList) + (RTShift*(dataTracker-1))
else:
	rawChart.x_axis.scaling.min = min(minRTList)
	rawChart.x_axis.scaling.max = max(maxRTList)

#Sets the min and max values for the x-axis in the normalized chart.
normChart.x_axis.scaling.min = min(minRTList)
normChart.x_axis.scaling.max = max(maxRTList)

#Sets the min and max values for the y-axis in the normalized chart.
normChart.y_axis.scaling.min = 0
normChart.y_axis.scaling.max = 100

#Sets the y-axis major units on the normalized chart.
normChart.y_axis.majorUnit = 100

#Sets the text style for the legend in both charts.
rawChart.legend.txPr = finalProp
normChart.legend.txPr = finalProp

#Saves the output Excel file.
outFile.save("Superimposed Plots.xlsx")
    
#Lets user know that the program is finished.
finished = raw_input("SUPER_ATM is finished! Press 'enter' to close the program.")